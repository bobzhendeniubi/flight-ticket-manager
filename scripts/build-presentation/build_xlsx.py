#!/usr/bin/env python3
"""Build docs/presentation/week-2/COSTS.xlsx — 4 tabs.
- Tab 1: 已花成本（开发期）
- Tab 2: 月成本预估（3 阶段）
- Tab 3: AI Token 成本拆解
- Tab 4: 单位经济学
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "/Users/bobwang/Documents/Flight Ticket Manager/docs/presentation/week-2/COSTS.xlsx"

BLUE = Font(name="Arial", color="0000FF")
BLACK = Font(name="Arial", color="000000")
GREEN = Font(name="Arial", color="008000")
HEADER = Font(name="Arial", color="FFFFFF", bold=True)
HEADER_FILL = PatternFill("solid", start_color="1E2761")
TOTAL_FILL = PatternFill("solid", start_color="FFF2CC")
SECTION_FILL = PatternFill("solid", start_color="E7E8D1")
THIN = Side(border_style="thin", color="CCCCCC")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

CURRENCY = '"$"#,##0;("$"#,##0);"-"'
CURRENCY_2DP = '"$"#,##0.00;("$"#,##0.00);"-"'
PCT = "0.0%"

wb = Workbook()

# ───────────────────────── Tab 1: 已花成本 ─────────────────────────
ws = wb.active
ws.title = "已花成本"
ws.sheet_view.showGridLines = False

ws["A1"] = "开发期已花成本（截至 2026-04-29）"
ws["A1"].font = Font(name="Arial", size=14, bold=True)
ws.merge_cells("A1:C1")

for i, h in enumerate(["类别", "金额（USD）", "说明"], 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = HEADER; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BOX

rows = [
    ("AI / LLM API（开发期）",   3200, "4 个月 AI 助手、Codex review、测试代码生成、提示工程迭代"),
    ("AI 编码工具订阅",          500, "Cursor / Claude Code / Codex 订阅"),
    ("服务器（阿里云 HK）",      200, "4C8G 轻量实例 包月 ~$32 × 6 期"),
    ("域名 + DNS",                50, "备用域名 + Cloudflare 免费层试用"),
    ("GitHub / CI",                80, "Actions 分钟、Codespaces 偶尔用"),
    ("OpenAI / 模型对比测试",    250, "早期对比 Claude vs GPT 选模型"),
    ("杂项",                      120, "图床（Unsplash CDN）、字体、人工核对越南语"),
    ("图标 / 设计素材",            80, "Figma 订阅 + 一些图标"),
    ("预留 buffer",               520, "重启 / 调错 / 数据库迁移期间小额浪费"),
]
sr = 4
for r, (cat, amt, note) in enumerate(rows, start=sr):
    ws.cell(row=r, column=1, value=cat).alignment = LEFT
    c = ws.cell(row=r, column=2, value=amt)
    c.font = BLUE; c.number_format = CURRENCY; c.alignment = RIGHT
    ws.cell(row=r, column=3, value=note).alignment = LEFT
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = BOX

tr = sr + len(rows)
ws.cell(row=tr, column=1, value="合计（已花）").font = Font(name="Arial", bold=True)
ws.cell(row=tr, column=1).fill = TOTAL_FILL
tc = ws.cell(row=tr, column=2, value=f"=SUM(B{sr}:B{tr-1})")
tc.font = Font(name="Arial", bold=True); tc.number_format = CURRENCY
tc.fill = TOTAL_FILL; tc.alignment = RIGHT
ws.cell(row=tr, column=3).fill = TOTAL_FILL
for col in range(1, 4):
    ws.cell(row=tr, column=col).border = BOX

ar = tr + 2
ws.cell(row=ar, column=1, value="AI 占比（API + 工具 + 测试）").font = Font(name="Arial", italic=True)
ws.cell(row=ar, column=2, value=f"=(B{sr}+B{sr+1}+B{sr+5})/B{tr}").number_format = PCT
ws.cell(row=ar, column=2).alignment = RIGHT

ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 60

# ───────────────────────── Tab 2: 月成本预估 ─────────────────────────
ws2 = wb.create_sheet("月成本预估")
ws2.sheet_view.showGridLines = False
ws2["A1"] = "投产后月成本预估（3 阶段）"
ws2["A1"].font = Font(name="Arial", size=14, bold=True)
ws2.merge_cells("A1:E1")

ws2["A3"] = "蓝色 = 假设输入；黑色 = 公式合计；金额单位：USD/月"
ws2["A3"].font = Font(name="Arial", italic=True, size=10, color="666666")
ws2.merge_cells("A3:E3")

for i, h in enumerate(["类别", "A · 内部测试期", "B · 公测期 (~500 客户)", "C · 稳定期 (~5000 客户)", "说明"], 1):
    c = ws2.cell(row=5, column=i, value=h)
    c.font = HEADER; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BOX

cost_rows = [
    ("AI / Token",            35,    550,    3500,  "对话量×单价；A: 50-200/天 ｜ B: 500-1500/天 ｜ C: 3000-8000/天"),
    ("服务器",                32,    100,    250,   "阿里云 HK 4C8G → 8C16G + 备份机 → 主从 + worker"),
    ("数据库（独立 RDS）",      0,    50,     150,   "B 起独立 4G；C 主从 + 自动备份"),
    ("CDN（Cloudflare）",       0,    20,     100,   "免费 → Pro $20 → Business $200，按流量分级"),
    ("邮件 / 短信",              0,    100,    600,   "阿里云通信 / Twilio 按量计费"),
    ("监控（Sentry + 日志）",    0,    30,     100,   "Team plan + 日志聚合"),
    ("其他（域名/SSL/备份）",    5,    20,     50,    "域名续费、SSL 免费、备份外发"),
]
st = 6
for r, (cat, a, b, c, note) in enumerate(cost_rows, start=st):
    ws2.cell(row=r, column=1, value=cat).alignment = LEFT
    for col, val in [(2, a), (3, b), (4, c)]:
        cell = ws2.cell(row=r, column=col, value=val)
        cell.font = BLUE; cell.number_format = CURRENCY; cell.alignment = RIGHT
    ws2.cell(row=r, column=5, value=note).alignment = LEFT
    for col in range(1, 6):
        ws2.cell(row=r, column=col).border = BOX

tr2 = st + len(cost_rows)
ws2.cell(row=tr2, column=1, value="月成本合计").font = Font(name="Arial", bold=True)
for col in range(1, 6):
    ws2.cell(row=tr2, column=col).fill = TOTAL_FILL
    ws2.cell(row=tr2, column=col).border = BOX
for col_letter in ("B", "C", "D"):
    cell = ws2[f"{col_letter}{tr2}"]
    cell.value = f"=SUM({col_letter}{st}:{col_letter}{tr2-1})"
    cell.font = Font(name="Arial", bold=True); cell.number_format = CURRENCY; cell.alignment = RIGHT

yr = tr2 + 1
ws2.cell(row=yr, column=1, value="× 12 月（年）").font = Font(name="Arial", italic=True)
for col_letter in ("B", "C", "D"):
    cell = ws2[f"{col_letter}{yr}"]
    cell.value = f"={col_letter}{tr2}*12"
    cell.number_format = CURRENCY; cell.alignment = RIGHT

ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 24
ws2.column_dimensions["E"].width = 50

# ───────────────────────── Tab 3: AI Token 成本 ─────────────────────────
ws3 = wb.create_sheet("AI Token 成本")
ws3.sheet_view.showGridLines = False
ws3["A1"] = "AI 单次对话成本（Claude Sonnet 4.6 定价）"
ws3["A1"].font = Font(name="Arial", size=14, bold=True)
ws3.merge_cells("A1:D1")

ws3["A3"] = "定价（USD per 1M tokens）"
ws3["A3"].font = Font(name="Arial", bold=True); ws3["A3"].fill = SECTION_FILL
ws3.merge_cells("A3:B3")

pricing = [("输入 token", 3.00), ("输出 token", 15.00), ("缓存读取（输入打 9 折）", 0.30)]
for r, (k, v) in enumerate(pricing, start=4):
    ws3.cell(row=r, column=1, value=k).alignment = LEFT
    c = ws3.cell(row=r, column=2, value=v)
    c.font = BLUE; c.number_format = CURRENCY_2DP; c.alignment = RIGHT
    for col in range(1, 3):
        ws3.cell(row=r, column=col).border = BOX

ws3["A8"] = "单次对话假设（5 个轮次平均）"
ws3["A8"].font = Font(name="Arial", bold=True); ws3["A8"].fill = SECTION_FILL
ws3.merge_cells("A8:D8")

assumps = [
    ("输入 tokens（无缓存）",      10000),
    ("输入 tokens（缓存命中）",     9000),
    ("输入 tokens（缓存未命中）",   1000),
    ("输出 tokens",                 2500),
]
for r, (k, v) in enumerate(assumps, start=9):
    ws3.cell(row=r, column=1, value=k).alignment = LEFT
    c = ws3.cell(row=r, column=2, value=v)
    c.font = BLUE; c.number_format = "#,##0"; c.alignment = RIGHT
    for col in range(1, 3):
        ws3.cell(row=r, column=col).border = BOX

ws3["A14"] = "成本计算"
ws3["A14"].font = Font(name="Arial", bold=True); ws3["A14"].fill = SECTION_FILL
ws3.merge_cells("A14:D14")

ws3["A15"] = "不带 prompt caching"
ws3["A15"].font = Font(name="Arial", italic=True)
ws3["A16"] = "输入成本"
ws3["B16"] = "=B9*B4/1000000"
ws3["B16"].number_format = CURRENCY_2DP; ws3["B16"].alignment = RIGHT
ws3["A17"] = "输出成本"
ws3["B17"] = "=B12*B5/1000000"
ws3["B17"].number_format = CURRENCY_2DP; ws3["B17"].alignment = RIGHT
ws3["A18"] = "对话总成本（无缓存）"
ws3["A18"].font = Font(name="Arial", bold=True)
ws3["B18"] = "=B16+B17"
ws3["B18"].number_format = CURRENCY_2DP
ws3["B18"].font = Font(name="Arial", bold=True); ws3["B18"].fill = TOTAL_FILL; ws3["B18"].alignment = RIGHT

ws3["A20"] = "启用 prompt caching"
ws3["A20"].font = Font(name="Arial", italic=True)
ws3["A21"] = "缓存读取成本"
ws3["B21"] = "=B10*B6/1000000"
ws3["B21"].number_format = CURRENCY_2DP; ws3["B21"].alignment = RIGHT
ws3["A22"] = "全价输入成本"
ws3["B22"] = "=B11*B4/1000000"
ws3["B22"].number_format = CURRENCY_2DP; ws3["B22"].alignment = RIGHT
ws3["A23"] = "输出成本"
ws3["B23"] = "=B12*B5/1000000"
ws3["B23"].number_format = CURRENCY_2DP; ws3["B23"].alignment = RIGHT
ws3["A24"] = "对话总成本（含缓存）"
ws3["A24"].font = Font(name="Arial", bold=True)
ws3["B24"] = "=B21+B22+B23"
ws3["B24"].number_format = CURRENCY_2DP
ws3["B24"].font = Font(name="Arial", bold=True); ws3["B24"].fill = TOTAL_FILL; ws3["B24"].alignment = RIGHT
ws3["A25"] = "省下"
ws3["B25"] = "=1-B24/B18"
ws3["B25"].number_format = PCT; ws3["B25"].alignment = RIGHT

ws3["A27"] = "对话量 → 月成本预估（含缓存）"
ws3["A27"].font = Font(name="Arial", bold=True); ws3["A27"].fill = SECTION_FILL
ws3.merge_cells("A27:D27")

for i, h in enumerate(["对话/天", "对话/月", "月成本（USD）", "年成本（USD）"], 1):
    c = ws3.cell(row=28, column=i, value=h)
    c.font = HEADER; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BOX

scenarios = [50, 200, 500, 1000, 2000, 5000, 8000]
for i, daily in enumerate(scenarios, start=29):
    ws3.cell(row=i, column=1, value=daily).font = BLUE
    ws3.cell(row=i, column=1).number_format = "#,##0"
    ws3.cell(row=i, column=1).alignment = RIGHT
    ws3.cell(row=i, column=2, value=f"=A{i}*30").number_format = "#,##0"
    ws3.cell(row=i, column=2).alignment = RIGHT
    ws3.cell(row=i, column=3, value=f"=B{i}*$B$24").number_format = CURRENCY
    ws3.cell(row=i, column=3).alignment = RIGHT
    ws3.cell(row=i, column=4, value=f"=C{i}*12").number_format = CURRENCY
    ws3.cell(row=i, column=4).alignment = RIGHT
    for col in range(1, 5):
        ws3.cell(row=i, column=col).border = BOX

ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 18

# ───────────────────────── Tab 4: 单位经济学 ─────────────────────────
ws4 = wb.create_sheet("单位经济学")
ws4.sheet_view.showGridLines = False
ws4["A1"] = "单位经济学（成本 vs 收入）"
ws4["A1"].font = Font(name="Arial", size=14, bold=True)
ws4.merge_cells("A1:F1")

ws4["A3"] = "蓝色 = 假设；黑色 = 公式；绿色 = 跨表引用"
ws4["A3"].font = Font(name="Arial", italic=True, size=10, color="666666")
ws4.merge_cells("A3:F3")

ws4["A5"] = "汇率 CNY → USD"
ws4["B5"] = 0.14; ws4["B5"].font = BLUE; ws4["B5"].number_format = "0.00"
ws4["A6"] = "平台抽成率"
ws4["B6"] = 0.08; ws4["B6"].font = BLUE; ws4["B6"].number_format = PCT

for i, h in enumerate(["阶段", "月单量", "客单价（¥）", "月 GMV（¥）", "月毛利（¥）", "月毛利（$）"], 1):
    c = ws4.cell(row=8, column=i, value=h)
    c.font = HEADER; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BOX

ue = [("公测期", 100, 3500), ("稳定期", 1000, 3800), ("规模化", 5000, 4000)]
for r, (stage, qty, aov) in enumerate(ue, start=9):
    ws4.cell(row=r, column=1, value=stage).alignment = LEFT
    c = ws4.cell(row=r, column=2, value=qty); c.font = BLUE; c.number_format = "#,##0"; c.alignment = RIGHT
    c = ws4.cell(row=r, column=3, value=aov); c.font = BLUE; c.number_format = "#,##0"; c.alignment = RIGHT
    ws4.cell(row=r, column=4, value=f"=B{r}*C{r}").number_format = "¥#,##0"
    ws4.cell(row=r, column=4).alignment = RIGHT
    ws4.cell(row=r, column=5, value=f"=D{r}*$B$6").number_format = "¥#,##0"
    ws4.cell(row=r, column=5).alignment = RIGHT
    ws4.cell(row=r, column=6, value=f"=E{r}*$B$5").number_format = CURRENCY
    ws4.cell(row=r, column=6).alignment = RIGHT
    for col in range(1, 7):
        ws4.cell(row=r, column=col).border = BOX

ws4["A14"] = "成本覆盖比"
ws4["A14"].font = Font(name="Arial", bold=True); ws4["A14"].fill = SECTION_FILL
ws4.merge_cells("A14:F14")

for i, h in enumerate(["阶段", "月毛利（$）", "月成本（$）", "覆盖倍数", "毛利率（毛利/GMV）", ""], 1):
    c = ws4.cell(row=15, column=i, value=h)
    c.font = HEADER; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = BOX

cov = [
    ("公测期", "F9",  "='月成本预估'!C13"),
    ("稳定期", "F10", "='月成本预估'!D13"),
    ("规模化", "F11", "='月成本预估'!D13*1.5"),
]
for r, (stage, profit_ref, cost_ref) in enumerate(cov, start=16):
    ws4.cell(row=r, column=1, value=stage).alignment = LEFT
    cell = ws4.cell(row=r, column=2, value=f"={profit_ref}")
    cell.number_format = CURRENCY; cell.font = GREEN; cell.alignment = RIGHT
    cell = ws4.cell(row=r, column=3, value=cost_ref)
    cell.number_format = CURRENCY; cell.font = GREEN; cell.alignment = RIGHT
    cell = ws4.cell(row=r, column=4, value=f"=B{r}/C{r}")
    cell.number_format = '0.0"×"'; cell.alignment = RIGHT
    src = {"公测期": 9, "稳定期": 10, "规模化": 11}[stage]
    cell = ws4.cell(row=r, column=5, value=f"=E{src}/D{src}")
    cell.number_format = PCT; cell.alignment = RIGHT
    for col in range(1, 6):
        ws4.cell(row=r, column=col).border = BOX

ws4.column_dimensions["A"].width = 14
for col_letter in ("B", "C", "D", "E", "F"):
    ws4.column_dimensions[col_letter].width = 18

ws4["A20"] = "结论：单位经济学健康。关键不是控成本，是订单量做起来。"
ws4["A20"].font = Font(name="Arial", italic=True, bold=True, color="2C5F2D")
ws4.merge_cells("A20:F20")

wb.save(OUT)
print(f"Saved: {OUT}")
