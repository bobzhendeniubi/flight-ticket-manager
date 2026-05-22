#!/usr/bin/env python3
"""Read docs/finances/COSTS.xlsx, emit backend/src/modules/finances/costs-data.ts

每次更新账本后跑：
  python3 scripts/build-presentation/build_costs_json.py
然后 commit + redeploy backend，admin 后台「财务」页就更新。
"""
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / 'docs' / 'finances' / 'COSTS.xlsx'
OUT  = ROOT / 'backend' / 'src' / 'modules' / 'finances' / 'costs-data.ts'

wb = load_workbook(XLSX, data_only=True)

# ── 1. 已花成本（汇总）──
ws = wb['已花成本']
summary_categories = []
summary_total = 0
title = ws['A1'].value or '开发期已花成本'
for r in range(4, ws.max_row + 1):
    label = ws.cell(row=r, column=1).value
    amt = ws.cell(row=r, column=2).value
    note = ws.cell(row=r, column=3).value or ''
    if not label:
        continue
    label_s = str(label).strip()
    if '合计' in label_s:
        summary_total = float(amt) if amt else 0
        continue
    if isinstance(amt, (int, float)):
        summary_categories.append({'label': label_s, 'usd': float(amt), 'note': str(note)})

# ── 2. 支出明细 ──
ws = wb['支出明细']
detail_rows = []
detail_total = 0
for r in range(5, ws.max_row + 1):
    a = ws.cell(row=r, column=1).value
    b = ws.cell(row=r, column=2).value
    c = ws.cell(row=r, column=3).value
    d = ws.cell(row=r, column=4).value
    e = ws.cell(row=r, column=5).value
    f = ws.cell(row=r, column=6).value
    if not a:
        continue
    a_s = str(a).strip()
    if '═' in a_s or (b is None and c is None and d is None):
        detail_rows.append({'isSection': True, 'label': a_s})
        continue
    if '合计' in a_s and b is None:
        detail_total = float(d) if d else 0
        continue
    detail_rows.append({
        'isSection': False,
        'date':     a_s,
        'category': str(b) if b else '',
        'vendor':   str(c) if c else '',
        'usd':      float(d) if isinstance(d, (int, float)) else 0,
        'what':     str(e) if e else '',
        'hours':    str(f) if f else '',
    })

# ── 3. 月成本预估 ──
ws = wb['月成本预估']
monthly = {'rows': [], 'totals': {}}
for r in range(6, ws.max_row + 1):
    label = ws.cell(row=r, column=1).value
    if not label:
        continue
    a_test = ws.cell(row=r, column=2).value
    b_beta = ws.cell(row=r, column=3).value
    c_stab = ws.cell(row=r, column=4).value
    note = ws.cell(row=r, column=5).value or ''
    if '月成本合计' in str(label):
        monthly['totals'] = {'testing': float(a_test or 0), 'beta': float(b_beta or 0), 'stable': float(c_stab or 0)}
        continue
    if '12' in str(label) and '月' in str(label):
        continue
    if isinstance(a_test, (int, float)):
        monthly['rows'].append({
            'category': str(label),
            'testing': float(a_test),
            'beta':    float(b_beta or 0),
            'stable':  float(c_stab or 0),
            'note':    str(note),
        })

# ── 4. 单位经济学 ──
ws = wb['单位经济学']
unit_econ = {'stages': []}
for r in range(9, ws.max_row + 1):
    stage = ws.cell(row=r, column=1).value
    qty = ws.cell(row=r, column=2).value
    aov = ws.cell(row=r, column=3).value
    gmv = ws.cell(row=r, column=4).value
    profit_cny = ws.cell(row=r, column=5).value
    profit_usd = ws.cell(row=r, column=6).value
    if not stage or '阶段' in str(stage) or '成本覆盖比' in str(stage):
        continue
    if '月毛利' in str(stage):
        break
    if isinstance(qty, (int, float)):
        unit_econ['stages'].append({
            'stage':       str(stage),
            'orders':      int(qty),
            'aovCny':      float(aov or 0),
            'gmvCny':      float(gmv or 0),
            'profitCny':   float(profit_cny or 0),
            'profitUsd':   float(profit_usd or 0),
        })

data = {
    'asOf':       '2026-05-20',
    'title':      title,
    'totalUsd':   summary_total,
    'categories': summary_categories,
    'detail': {
        'rows':     detail_rows,
        'totalUsd': detail_total,
    },
    'monthly':  monthly,
    'unitEcon': unit_econ,
}

ts_content = f"""// AUTO-GENERATED from docs/finances/COSTS.xlsx by scripts/build-presentation/build_costs_json.py
// 修改 COSTS.xlsx 后跑：
//   python3 scripts/build-presentation/build_costs_json.py
// 然后 commit + redeploy backend。

export interface CostCategory {{ label: string; usd: number; note: string; }}
export interface CostDetailRow {{
  isSection: boolean;
  label?: string;
  date?: string;
  category?: string;
  vendor?: string;
  usd?: number;
  what?: string;
  hours?: string;
}}
export interface MonthlyForecastRow {{
  category: string;
  testing: number;
  beta: number;
  stable: number;
  note: string;
}}
export interface UnitEconStage {{
  stage: string;
  orders: number;
  aovCny: number;
  gmvCny: number;
  profitCny: number;
  profitUsd: number;
}}

export interface CostsData {{
  asOf: string;
  title: string;
  totalUsd: number;
  categories: CostCategory[];
  detail: {{ rows: CostDetailRow[]; totalUsd: number; }};
  monthly: {{
    rows: MonthlyForecastRow[];
    totals: {{ testing: number; beta: number; stable: number; }};
  }};
  unitEcon: {{ stages: UnitEconStage[]; }};
}}

export const COSTS_DATA: CostsData = {json.dumps(data, ensure_ascii=False, indent=2)} as const;
"""

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(ts_content, encoding='utf-8')
print(f'✓ {OUT}')
print(f'  total: ${summary_total:,.0f}')
print(f'  categories: {len(summary_categories)}')
print(f'  detail rows: {len(detail_rows)}')
print(f'  monthly forecast rows: {len(monthly["rows"])}')
print(f'  unit econ stages: {len(unit_econ["stages"])}')
